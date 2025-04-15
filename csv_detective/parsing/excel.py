from io import BytesIO
from time import time
from typing import Optional

import openpyxl
import pandas as pd
import requests
import xlrd

from csv_detective.detection.engine import engine_to_file
from csv_detective.detection.rows import remove_empty_first_rows
from csv_detective.utils import (
    display_logs_depending_process_time,
    is_url,
)

NEW_EXCEL_EXT = [".xlsx", ".xlsm", ".xltx", ".xltm"]
OLD_EXCEL_EXT = [".xls"]
OPEN_OFFICE_EXT = [".odf", ".ods", ".odt"]
XLS_LIKE_EXT = NEW_EXCEL_EXT + OLD_EXCEL_EXT + OPEN_OFFICE_EXT


def parse_excel(
    file_path: str,
    num_rows: int = -1,
    engine: Optional[str] = None,
    sheet_name: Optional[str] = None,
    random_state: int = 42,
    verbose: bool = False,
) -> tuple[pd.DataFrame, int, int, str, str, int]:
    """"Excel-like parsing is really slow, could be a good improvement for future development"""
    if verbose:
        start = time()
    no_sheet_specified = sheet_name is None

    if (
        engine in ['openpyxl', 'xlrd'] or
        any([file_path.endswith(k) for k in NEW_EXCEL_EXT + OLD_EXCEL_EXT])
    ):
        remote_content = None
        if is_url(file_path):
            r = requests.get(file_path)
            r.raise_for_status()
            remote_content = BytesIO(r.content)
        if not engine:
            if any([file_path.endswith(k) for k in NEW_EXCEL_EXT]):
                engine = "openpyxl"
            else:
                engine = "xlrd"
        if sheet_name is None:
            if verbose:
                display_logs_depending_process_time(
                    f'Detected {engine_to_file[engine]} file, no sheet specified, reading the largest one',
                    time() - start,
                )
            try:
                if engine == "openpyxl":
                    # openpyxl doesn't want to open files that don't have a valid extension
                    # see: https://foss.heptapod.net/openpyxl/openpyxl/-/issues/2157
                    # if the file is remote, we have a remote content anyway so it's fine
                    if not remote_content and '.' not in file_path.split('/')[-1]:
                        with open(file_path, 'rb') as f:
                            remote_content = BytesIO(f.read())
                    # faster than loading all sheets
                    wb = openpyxl.load_workbook(remote_content or file_path, read_only=True)
                    try:
                        sizes = {s.title: s.max_row * s.max_column for s in wb.worksheets}
                    except TypeError:
                        # sometimes read_only can't get the info, so we have to open the file for real
                        # this takes more time but it's for a limited number of files
                        # and it's this or nothing
                        wb = openpyxl.load_workbook(remote_content or file_path)
                        sizes = {s.title: s.max_row * s.max_column for s in wb.worksheets}
                else:
                    if remote_content:
                        wb = xlrd.open_workbook(file_contents=remote_content.read())
                    else:
                        wb = xlrd.open_workbook(file_path)
                    sizes = {s.name: s.nrows * s.ncols for s in wb.sheets()}
                sheet_name = max(sizes, key=sizes.get)
            except xlrd.biffh.XLRDError:
                # sometimes a xls file is recognized as ods
                if verbose:
                    display_logs_depending_process_time(
                        'Could not read file with classic xls reader, trying with ODS',
                        time() - start,
                    )
                engine = "odf"

    if engine == "odf" or any([file_path.endswith(k) for k in OPEN_OFFICE_EXT]):
        # for ODS files, no way to get sheets' sizes without
        # loading the file one way or another (pandas or pure odfpy)
        # so all in one
        engine = "odf"
        if sheet_name is None:
            if verbose:
                display_logs_depending_process_time(
                    f'Detected {engine_to_file[engine]} file, no sheet specified, reading the largest one',
                    time() - start,
                )
            tables = pd.read_excel(
                file_path,
                engine="odf",
                sheet_name=None,
                dtype="unicode",
            )
            sizes = {sheet_name: table.size for sheet_name, table in tables.items()}
            sheet_name = max(sizes, key=sizes.get)
            if verbose:
                display_logs_depending_process_time(
                    f'Going forwards with sheet "{sheet_name}"',
                    time() - start,
                )
            table = tables[sheet_name]
        else:
            if verbose:
                display_logs_depending_process_time(
                    f'Detected {engine_to_file[engine]} file, reading sheet "{sheet_name}"',
                    time() - start,
                )
            table = pd.read_excel(
                file_path,
                engine="odf",
                sheet_name=sheet_name,
                dtype="unicode",
            )
        table, header_row_idx = remove_empty_first_rows(table)
        total_lines = len(table)
        nb_duplicates = len(table.loc[table.duplicated()])
        if num_rows > 0:
            num_rows = min(num_rows - 1, total_lines)
            table = table.sample(num_rows, random_state=random_state)
        if verbose:
            display_logs_depending_process_time(
                f'Table parsed successfully in {round(time() - start, 3)}s',
                time() - start,
            )
        return table, total_lines, nb_duplicates, sheet_name, engine, header_row_idx

    # so here we end up with (old and new) excel files only
    if verbose:
        if no_sheet_specified:
            display_logs_depending_process_time(
                f'Going forwards with sheet "{sheet_name}"',
                time() - start,
            )
        else:
            display_logs_depending_process_time(
                f'Detected {engine_to_file[engine]} file, reading sheet "{sheet_name}"',
                time() - start,
            )
    table = pd.read_excel(
        file_path,
        engine=engine,
        sheet_name=sheet_name,
        dtype="unicode",
    )
    table, header_row_idx = remove_empty_first_rows(table)
    total_lines = len(table)
    nb_duplicates = len(table.loc[table.duplicated()])
    if num_rows > 0:
        num_rows = min(num_rows - 1, total_lines)
        table = table.sample(num_rows, random_state=random_state)
    if verbose:
        display_logs_depending_process_time(
            f'Table parsed successfully in {round(time() - start, 3)}s',
            time() - start,
        )
    return table, total_lines, nb_duplicates, sheet_name, engine, header_row_idx
