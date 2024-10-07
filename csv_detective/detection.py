from typing import TextIO, Optional
import pandas as pd
import math
import csv
from cchardet import detect
from ast import literal_eval
import logging
from time import time
import openpyxl
import xlrd
import requests
from io import BytesIO
import magic
from csv_detective.utils import display_logs_depending_process_time
from csv_detective.detect_fields.other.float import float_casting

logging.basicConfig(level=logging.INFO)

NEW_EXCEL_EXT = [".xlsx", ".xlsm", ".xltx", ".xltm"]
OLD_EXCEL_EXT = [".xls"]
OPEN_OFFICE_EXT = [".odf", ".ods", ".odt"]
XLS_LIKE_EXT = NEW_EXCEL_EXT + OLD_EXCEL_EXT + OPEN_OFFICE_EXT
engine_to_file = {
    "openpyxl": "Excel",
    "xlrd": "old Excel",
    "odf": "OpenOffice"
}


def is_url(csv_file_path: str):
    # could be more sophisticated if needed
    return csv_file_path.startswith('http')


def detect_continuous_variable(table: pd.DataFrame, continuous_th: float = 0.9, verbose: bool = False):
    """
    Detects whether a column contains continuous variables. We consider a continuous column
    one that contains
    a considerable amount of float values.
    We removed the integers as we then end up with postal codes, insee codes, and all sort
    of codes and types.
    This is not optimal but it will do for now.
    :param table:
    :return:
    """
    # if we need this again in the future, could be first based on columns detected as int/float to cut time

    def check_threshold(serie: pd.Series, continuous_th: float):
        count = serie.value_counts().to_dict()
        total_nb = len(serie)
        if float in count:
            nb_floats = count[float]
        else:
            return False
        if nb_floats / total_nb >= continuous_th:
            return True
        else:
            return False

    def parses_to_integer(value: str):
        try:
            value = value.replace(",", ".")
            value = literal_eval(value)
            return type(value)
        # flake8: noqa
        except:
            return False

    if verbose:
        start = time()
        logging.info("Detecting continuous columns")
    res = table.apply(
        lambda serie: check_threshold(serie.apply(parses_to_integer), continuous_th)
    )
    if verbose:
        display_logs_depending_process_time(
            f"Detected {sum(res)} continuous columns in {round(time() - start, 3)}s",
            time() - start
        )
    return res.index[res]


def detetect_categorical_variable(
    table: pd.DataFrame,
    threshold_pct_categorical: float = 0.05,
    max_number_categorical_values: int = 25,
    verbose: bool = False,
):
    """
    Heuristically detects whether a table (df) contains categorical values according to
    the number of unique values contained.
    As the idea of detecting categorical values is to then try to learn models to predict
    them, we limit categorical values to at most 25 different modes or at most 5% disparity.
    Postal code, insee code, code region and so on, may be thus not considered categorical values.
    :param table:
    :param threshold_pct_categorical:
    :param max_number_categorical_values:
    :return:
    """

    def abs_number_different_values(column_values: pd.Series):
        return column_values.nunique()

    def rel_number_different_values(column_values: pd.Series):
        return column_values.nunique() / len(column_values)

    def detect_categorical(column_values: pd.Series):
        abs_unique_values = abs_number_different_values(column_values)
        rel_unique_values = rel_number_different_values(column_values)
        if (
            abs_unique_values <= max_number_categorical_values
            or rel_unique_values <= threshold_pct_categorical
        ):
            return True
        return False

    if verbose:
        start = time()
        logging.info("Detecting categorical columns")
    res = table.apply(lambda serie: detect_categorical(serie))
    if verbose:
        display_logs_depending_process_time(
            f"Detected {sum(res)} categorical columns out of {len(table.columns)} in {round(time() - start, 3)}s",
            time() - start
        )
    return res.index[res], res


def detect_engine(csv_file_path: str, verbose=False):
    if verbose:
        start = time()
    mapping = {
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': 'openpyxl',
        'application/vnd.ms-excel': 'xlrd',
        'application/vnd.oasis.opendocument.spreadsheet': 'odf',
        # all these files could be recognized as zip, may need to check all cases then
        'application/zip': 'openpyxl',
    }
    # if none of the above, we move forwards with the csv process
    if is_url(csv_file_path):
        remote_content = requests.get(csv_file_path).content
        engine = mapping.get(magic.from_buffer(remote_content, mime=True))
    else:
        engine = mapping.get(magic.from_file(csv_file_path, mime=True))
    if verbose:
        display_logs_depending_process_time(
            f'File has no extension, detected {engine_to_file.get(engine, "csv")}',
            time() - start
        )
    return engine


def detect_separator(file: TextIO, verbose: bool = False):
    """Detects csv separator"""
    # TODO: add a robust detection:
    # si on a un point virgule comme texte et \t comme séparateur, on renvoie
    # pour l'instant un point virgule
    if verbose:
        start = time()
        logging.info("Detecting separator")
    file.seek(0)
    header = file.readline()
    possible_separators = [";", ",", "|", "\t"]
    sep_count = dict()
    for sep in possible_separators:
        sep_count[sep] = header.count(sep)
    sep = max(sep_count, key=sep_count.get)
    # testing that the first 10 (arbitrary) rows all have the same number of fields
    # as the header. Prevents downstream unwanted behaviour where pandas can load
    # the file (in a weird way) but the process is irrelevant.
    file.seek(0)
    reader = csv.reader(file, delimiter=sep)
    rows_lengths = set()
    for idx, row in enumerate(reader):
        if idx > 10:
            break
        rows_lengths.add(len(row))
    if len(rows_lengths) > 1:
        raise ValueError('Number of columns is not even across the first 10 rows.')

    if verbose:
        display_logs_depending_process_time(
            f'Detected separator: "{sep}" in {round(time() - start, 3)}s',
            time() - start
        )
    return sep


def detect_encoding(csv_file_path: str, verbose: bool = False):
    """
    Detects file encoding using faust-cchardet (forked from the original cchardet)
    """
    if verbose:
        start = time()
        logging.info("Detecting encoding")
    if is_url(csv_file_path):
        r = requests.get(csv_file_path)
        r.raise_for_status()
        binary_file = BytesIO(r.content)
    else:
        binary_file = open(csv_file_path, mode="rb")
    encoding_dict = detect(binary_file.read())
    if verbose:
        message = f'Detected encoding: "{encoding_dict["encoding"]}"'
        message += f' in {round(time() - start, 3)}s (confidence: {round(encoding_dict["confidence"]*100)}%)'
        display_logs_depending_process_time(
            message,
            time() - start
        )
    return encoding_dict['encoding']


def parse_table(
    the_file: TextIO,
    encoding: str,
    sep: str,
    num_rows: int,
    skiprows: int,
    random_state: int = 42,
    verbose : bool = False,
):
    # Takes care of some problems
    if verbose:
        start = time()
        logging.info("Parsing table")
    table = None

    if not isinstance(the_file, str):
        the_file.seek(0)

    total_lines = None
    for encoding in [encoding, "ISO-8859-1", "utf-8"]:
        # TODO : modification systematique
        if encoding is None:
            continue

        if "ISO-8859" in encoding:
            encoding = "ISO-8859-1"
        try:
            table = pd.read_csv(
                the_file, sep=sep, dtype="unicode", encoding=encoding, skiprows=skiprows
            )
            total_lines = len(table)
            nb_duplicates = len(table.loc[table.duplicated()])
            if num_rows > 0:
                num_rows = min(num_rows - 1, total_lines)
                table = table.sample(num_rows, random_state=random_state)
            # else : table is unchanged
            break
        except TypeError:
            print("Trying encoding : {encoding}".format(encoding=encoding))

    if table is None:
        logging.error("  >> encoding not found")
        return table, "NA", "NA"
    if verbose:
        display_logs_depending_process_time(
            f'Table parsed successfully in {round(time() - start, 3)}s',
            time() - start
        )
    return table, total_lines, nb_duplicates


def remove_empty_first_rows(table: pd.DataFrame):
    """Analog process to detect_headers for csv files, determines how many rows to skip
    to end up with the header at the right place"""
    idx = 0
    if all([str(c).startswith('Unnamed:') for c in table.columns]):
        # there is on offset between the index in the file (idx here)
        # and the index in the dataframe, because of the header
        idx = 1
        while table.iloc[idx - 1].isna().all():
            idx += 1
        cols = table.iloc[idx - 1]
        table = table.iloc[idx:]
        table.columns = cols.to_list()
    # +1 here because the columns should count as a row
    return table, idx


def parse_excel(
    csv_file_path: str,
    num_rows: int = -1,
    engine: Optional[str] = None,
    sheet_name: Optional[str] = None,
    random_state: int = 42,
    verbose : bool = False,
):
    """"Excel-like parsing is really slow, could be a good improvement for future development"""
    if verbose:
        start = time()
    no_sheet_specified = sheet_name is None

    if (
        engine in ['openpyxl', 'xlrd'] or
        any([csv_file_path.endswith(k) for k in NEW_EXCEL_EXT + OLD_EXCEL_EXT])
    ):
        remote_content = None
        if is_url(csv_file_path):
            r = requests.get(csv_file_path)
            r.raise_for_status()
            remote_content = BytesIO(r.content)
        if not engine:
            if any([csv_file_path.endswith(k) for k in NEW_EXCEL_EXT]):
                engine = "openpyxl"
            else:
                engine = "xlrd"
        if sheet_name is None:
            if verbose:
                display_logs_depending_process_time(
                    f'Detected {engine_to_file[engine]} file, no sheet specified, reading the largest one',
                    time() - start
                )
            try:
                if engine == "openpyxl":
                    # openpyxl doesn't want to open files that don't have a valid extension
                    # see: https://foss.heptapod.net/openpyxl/openpyxl/-/issues/2157
                    # if the file is remote, we have a remote content anyway so it's fine
                    if not remote_content and '.' not in csv_file_path.split('/')[-1]:
                        with open(csv_file_path, 'rb') as f:
                            remote_content = BytesIO(f.read())
                    # faster than loading all sheets
                    wb = openpyxl.load_workbook(remote_content or csv_file_path, read_only=True)
                    try:
                        sizes = {s.title: s.max_row * s.max_column for s in wb.worksheets}
                    except TypeError:
                        # sometimes read_only can't get the info, so we have to open the file for real
                        # this takes more time but it's for a limited number of files
                        # and it's this or nothing
                        wb = openpyxl.load_workbook(remote_content or csv_file_path)
                        sizes = {s.title: s.max_row * s.max_column for s in wb.worksheets}
                else:
                    if remote_content:
                        wb = xlrd.open_workbook(file_contents=remote_content.read())
                    else:
                        wb = xlrd.open_workbook(csv_file_path)
                    sizes = {s.name: s.nrows * s.ncols for s in wb.sheets()}
                sheet_name = max(sizes, key=sizes.get)
            except xlrd.biffh.XLRDError:
                # sometimes a xls file is recognized as ods
                if verbose:
                    display_logs_depending_process_time(
                        'Could not read file with classic xls reader, trying with ODS',
                        time() - start
                    )
                engine = "odf"

    if engine == "odf" or any([csv_file_path.endswith(k) for k in OPEN_OFFICE_EXT]):
        # for ODS files, no way to get sheets' sizes without
        # loading the file one way or another (pandas or pure odfpy)
        # so all in one
        engine = "odf"
        if sheet_name is None:
            if verbose:
                display_logs_depending_process_time(
                    f'Detected {engine_to_file[engine]} file, no sheet specified, reading the largest one',
                    time() - start
                )
            tables = pd.read_excel(
                csv_file_path,
                engine="odf",
                sheet_name=None,
                dtype="unicode"
            )
            sizes = {sheet_name: table.size for sheet_name, table in tables.items()}
            sheet_name = max(sizes, key=sizes.get)
            if verbose:
                display_logs_depending_process_time(
                    f'Going forwards with sheet "{sheet_name}"',
                    time() - start
                )
            table = tables[sheet_name]
        else:
            if verbose:
                display_logs_depending_process_time(
                    f'Detected {engine_to_file[engine]} file, reading sheet "{sheet_name}"',
                    time() - start
                )
            table = pd.read_excel(
                csv_file_path,
                engine="odf",
                sheet_name=sheet_name,
                dtype="unicode"
            )
        table, header_row_idx = remove_empty_first_rows(table)
        total_lines = len(table)
        nb_duplicates = len(table.loc[table.duplicated()])
        if num_rows > 0:
            num_rows = min(num_rows - 1, total_lines)
            table = table.sample(num_rows, random_state=random_state)
        if verbose:
            display_logs_depending_process_time(
                f'Table parsed successfully in {round(time() - start, 3)}s',
                time() - start
            )
        return table, total_lines, nb_duplicates, sheet_name, engine, header_row_idx

    # so here we end up with (old and new) excel files only
    if verbose:
        if no_sheet_specified:
            display_logs_depending_process_time(
                f'Going forwards with sheet "{sheet_name}"',
                time() - start
            )
        else:
            display_logs_depending_process_time(
                f'Detected {engine_to_file[engine]} file, reading sheet "{sheet_name}"',
                time() - start
            )
    table = pd.read_excel(
        csv_file_path,
        engine=engine,
        sheet_name=sheet_name,
        dtype="unicode"
    )
    table, header_row_idx = remove_empty_first_rows(table)
    total_lines = len(table)
    nb_duplicates = len(table.loc[table.duplicated()])
    if num_rows > 0:
        num_rows = min(num_rows - 1, total_lines)
        table = table.sample(num_rows, random_state=random_state)
    if verbose:
        display_logs_depending_process_time(
            f'Table parsed successfully in {round(time() - start, 3)}s',
            time() - start
        )
    return table, total_lines, nb_duplicates, sheet_name, engine, header_row_idx


def prevent_nan(value: float):
    if math.isnan(value):
        return None
    return value


def create_profile(
    table: pd.DataFrame,
    dict_cols_fields: dict,
    num_rows: int,
    limited_output: bool = True,
    verbose: bool = False,
):
    if verbose:
        start = time()
        logging.info("Creating profile")
    map_python_types = {
        "string": str,
        "int": float,
        "float": float,
    }

    if num_rows > 0:
        raise ValueError("To create profiles num_rows has to be set to -1")
    safe_table = table.copy()
    if not limited_output:
        dict_cols_fields = {
            k: v[0] if v else {'python_type': 'string', 'format': 'string', 'score': 1.0}
            for k, v in dict_cols_fields.items()
        }
    dtypes = {
        k: map_python_types.get(v["python_type"], str)
        for k, v in dict_cols_fields.items()
    }
    for c in safe_table.columns:
        if dtypes[c] == float:
            safe_table[c] = safe_table[c].apply(
                lambda s: float_casting(s) if isinstance(s, str) else s
            )
    profile = {}
    for c in safe_table.columns:
        profile[c] = {}
        if map_python_types.get(dict_cols_fields[c]["python_type"], str) in [
            float,
            int,
        ]:
            profile[c].update(
                min=prevent_nan(map_python_types.get(dict_cols_fields[c]["python_type"], str)(
                    safe_table[c].min()
                )),
                max=prevent_nan(map_python_types.get(dict_cols_fields[c]["python_type"], str)(
                    safe_table[c].max()
                )),
                mean=prevent_nan(map_python_types.get(dict_cols_fields[c]["python_type"], str)(
                    safe_table[c].mean()
                )),
                std=prevent_nan(map_python_types.get(dict_cols_fields[c]["python_type"], str)(
                    safe_table[c].std()
                )),
            )
        tops_bruts = safe_table[safe_table[c].notna()][c] \
                .value_counts(dropna=True) \
                .reset_index() \
                .iloc[:10] \
                .to_dict(orient="records")
        tops = []
        for tb in tops_bruts:
            top = {}
            top["count"] = tb["count"]
            top["value"] = tb[c]
            tops.append(top)
        profile[c].update(
            tops=tops,
            nb_distinct=safe_table[c].nunique(),
            nb_missing_values=len(safe_table[c].loc[safe_table[c].isna()]),
        )
    if verbose:
        display_logs_depending_process_time(
            f"Created profile in {round(time() - start, 3)}s",
            time() - start
        )
    return profile


def detect_extra_columns(file: TextIO, sep: str):
    """regarde s'il y a des colonnes en trop
    Attention, file ne doit pas avoir de ligne vide"""
    file.seek(0)
    retour = False
    nb_useless_col = 99999

    for i in range(10):
        line = file.readline()
        # regarde si on a un retour
        if retour:
            assert line[-1] == "\n"
        if line[-1] == "\n":
            retour = True

        # regarde le nombre de derniere colonne inutile
        deb = 0 + retour
        line = line[::-1][deb:]
        k = 0
        for sign in line:
            if sign != sep:
                break
            k += 1
        if k == 0:
            return 0, retour
        nb_useless_col = min(k, nb_useless_col)
    return nb_useless_col, retour


def detect_headers(file: TextIO, sep: str, verbose: bool = False):
    """Tests 10 first rows for possible header (header not in 1st line)"""
    if verbose:
        start = time()
        logging.info("Detecting headers")
    file.seek(0)
    for i in range(10):
        header = file.readline()
        position = file.tell()
        chaine = [c for c in header.replace("\n", "").split(sep) if c]
        if chaine[-1] not in ["", "\n"] and all(
            [mot not in ["", "\n"] for mot in chaine[1:-1]]
        ):
            next_row = file.readline()
            file.seek(position)
            if header != next_row:
                if verbose:
                    display_logs_depending_process_time(
                        f'Detected headers in {round(time() - start, 3)}s',
                        time() - start
                    )
                return i, chaine
    if verbose:
        logging.info(f'No header detected')
    return 0, None


def detect_heading_columns(file: TextIO, sep: str, verbose : bool = False):
    """Tests first 10 lines to see if there are empty heading columns"""
    if verbose:
        start = time()
        logging.info("Detecting heading columns")
    file.seek(0)
    return_int = float("Inf")
    for i in range(10):
        line = file.readline()
        return_int = min(return_int, len(line) - len(line.strip(sep)))
        if return_int == 0:
            if verbose:
                display_logs_depending_process_time(
                    f'No heading column detected in {round(time() - start, 3)}s',
                    time() - start
                )
            return 0
    if verbose:
        display_logs_depending_process_time(
            f'{return_int} heading columns detected in {round(time() - start, 3)}s',
            time() - start
        )
    return return_int


def detect_trailing_columns(file: TextIO, sep: str, heading_columns: int, verbose : bool = False):
    """Tests first 10 lines to see if there are empty trailing columns"""
    if verbose:
        start = time()
        logging.info("Detecting trailing columns")
    file.seek(0)
    return_int = float("Inf")
    for i in range(10):
        line = file.readline()
        return_int = min(
            return_int,
            len(line.replace("\n", ""))
            - len(line.replace("\n", "").strip(sep))
            - heading_columns,
        )
        if return_int == 0:
            if verbose:
                display_logs_depending_process_time(
                    f'No trailing column detected in {round(time() - start, 3)}s',
                    time() - start
                )
            return 0
    if verbose:
        display_logs_depending_process_time(
            f'{return_int} trailing columns detected in {round(time() - start, 3)}s',
            time() - start
        )
    return return_int
